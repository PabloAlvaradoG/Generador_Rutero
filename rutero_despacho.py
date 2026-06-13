# ============================================================
#  RUTERO DE DESPACHO — Streamlit App v2
#  Modos: Por Destino | Por Camión
# ============================================================

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, datetime
from collections import defaultdict

# ── Paleta ───────────────────────────────────────────────────
C_AZO = "2C5F8A"; C_AZM = "5B9BD5"; C_AZC = "DEEAF7"
C_BLN = "FFFFFF"; C_GRS = "F5F5F5"; C_BRD = "C0C0C0"
C_NAR = "ED7D31"; C_RJO = "C00000"; C_VRD = "375623"
C_AMR = "FFF2CC"; C_MNT = "D9EAD3"; C_SLM = "FCE4D6"

# ── Helpers estilo ────────────────────────────────────────────
def thin(color=C_BRD):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def med_bottom(color=C_AZO):
    t = Side(style="thin", color=C_BRD)
    m = Side(style="medium", color=color)
    return Border(left=t, right=t, top=t, bottom=m)

def sc(ws, r, c, v=None, bold=False, sz=9, fg="000000",
       bg=None, ha="center", fmt=None, italic=False):
    cell = ws.cell(r, c)
    if v is not None: cell.value = v
    cell.font = Font(name="Arial", bold=bold, size=sz, color=fg, italic=italic)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical="center")
    cell.border = thin()
    if fmt: cell.number_format = fmt
    return cell

def mc(ws, r1, c1, r2, c2, v=None, bold=False, sz=9,
       fg="000000", bg=None, ha="center", fmt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    if v is not None: cell.value = v
    cell.font = Font(name="Arial", bold=bold, size=sz, color=fg)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical="center")
    cell.border = thin()
    if fmt: cell.number_format = fmt
    return cell

def lbl(ws, r, c1, c2, txt):
    if c1 < c2: ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(r, c1)
    c.value = txt
    c.font  = Font(name="Arial", bold=True, size=8, color=C_AZO)
    c.fill  = PatternFill("solid", fgColor=C_BLN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin()

def val(ws, r, c1, c2, v, bold=False, fmt=None):
    if c1 < c2: ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(r, c1)
    c.value = v
    c.font  = Font(name="Arial", bold=bold, size=9, color="323232")
    c.fill  = PatternFill("solid", fgColor=C_BLN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin()
    if fmt: c.number_format = fmt

def linea(ws, r, c1, c2):
    if c1 < c2: ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(r, c1)
    c.fill = PatternFill("solid", fgColor=C_BLN)
    t = Side(style="thin", color=C_BRD)
    b = Side(style="thin", color="646464")
    c.border = Border(left=t, right=t, top=t, bottom=b)

# ── Utilidades ────────────────────────────────────────────────
def norm(v): return str(v).strip().upper() if v else ""
def norm_key(v): return str(v).strip().upper() if v else ""

def sh_name(txt):
    s = str(txt)[:28]
    for ch in "/\\:*?[]!": s = s.replace(ch, "-")
    return s

def fecha_celda(cell):
    v = cell.value
    if v is None: return 0, "--"
    if isinstance(v, (datetime.datetime, datetime.date)):
        if isinstance(v, datetime.date) and not isinstance(v, datetime.datetime):
            v = datetime.datetime(v.year, v.month, v.day)
        base  = datetime.datetime(1899, 12, 30)
        serial = (v - base).days
        return serial, v.strftime("%d/%m/%Y")
    if isinstance(v, (int, float)):
        s = int(v)
        if s > 60: s -= 1
        d = datetime.date(1900,1,1) + datetime.timedelta(days=s-1)
        return int(v), d.strftime("%d/%m/%Y")
    return 0, str(v)

# ── Maestros ──────────────────────────────────────────────────
ALIASES = {
    "fecha":   ["fecha","fecha factura","date"],
    "comp":    ["comprobante","no. comprobante","numero comprobante","nro comprobante"],
    "total":   ["total","monto","valor","importe"],
    "codcli":  ["codigo cliente","cod. cliente","cod cliente","id cliente"],
    "nomdest": ["nombre destinatario","destinatario","nombre cliente"],
    "ciudad":  ["ciudad destinatario","ciudad dest","ciudad","destino"],
    "ruta":    ["ruta","cod ruta","codigo ruta"],
    "bodega":  ["bodega","centro sap","cod bodega"],
}
DEFAULTS = {"fecha":1,"comp":4,"total":5,"codcli":6,
            "nomdest":9,"ciudad":10,"ruta":11,"bodega":17}

def match_alias(col_n, aliases):
    for a in aliases:
        if col_n == a: return True
    for a in aliases:
        if " " in a and a in col_n: return True
    return False

def detectar_cols(ws, hdr_row=2):
    ci = dict(DEFAULTS)
    for c in range(1, 31):
        v = ws.cell(hdr_row, c).value
        if not v: continue
        cn = str(v).strip().lower()
        for field, als in ALIASES.items():
            if match_alias(cn, als):
                ci[field] = c
    return ci

def cargar_centros(wb):
    if "CENTROS" not in wb.sheetnames:
        return {}, {}, "No se encontró la hoja CENTROS"
    ws = wb["CENTROS"]
    nom, emp = {}, {}
    for i in range(3, ws.max_row+1):
        v = ws.cell(i,1).value
        if not v: continue
        try:    key = str(int(float(str(v))))
        except: key = norm(str(v))
        if key and key not in nom:
            nom[key] = str(ws.cell(i,2).value or "").strip()
            emp[key] = str(ws.cell(i,3).value or "").strip()
    return nom, emp, None

def cargar_clientes(wb):
    """
    CLIENTES: cabecera fila 2, datos desde fila 3
    Cols: A=Código B=Nombre C=Agente D=Prioridad
    Retorna: dict cod->agente, dict cod->prioridad, duplicados[], error
    """
    if "CLIENTES" not in wb.sheetnames:
        return {}, {}, [], "No se encontró la hoja CLIENTES"
    ws = wb["CLIENTES"]
    cli_vend = {}; cli_prio = {}
    duplicados = []
    from collections import Counter
    codigos_vistos = Counter()

    for i in range(3, ws.max_row+1):
        v = ws.cell(i,1).value
        if not v: continue
        key = norm(str(v))
        codigos_vistos[key] += 1

    for i in range(3, ws.max_row+1):
        v = ws.cell(i,1).value
        if not v: continue
        key = norm(str(v))
        nom = str(ws.cell(i,2).value or "").strip()
        agente = str(ws.cell(i,3).value or "").strip()
        try:
            prio = int(float(str(ws.cell(i,4).value)))
            if prio < 1 or prio > 998: prio = 999
        except: prio = 999

        if codigos_vistos[key] > 1:
            duplicados.append({
                "codigo": key, "nombre": nom,
                "agente": agente, "prioridad": prio, "fila": i
            })

        if key not in cli_vend:
            cli_vend[key] = agente
            cli_prio[key] = prio

    return cli_vend, cli_prio, duplicados, None

def cargar_rutas(wb):
    """
    RUTAS: cabecera fila 2, datos desde fila 3
    Cols: A=Camión B=Ciudad C=Agente D=Día E=Fecha
    Retorna:
      - ruta_camion: ciudad -> camion
      - ruta_dia:    ciudad -> dia
      - ruta_fecha:  ciudad -> fecha
      - agente_camion: agente -> camion (para rutero por camión)
      - agente_dia:    agente -> dia
      - camion_ciudades: camion -> [ciudades]
    """
    if "RUTAS" not in wb.sheetnames:
        return {},{},{},{},{},{}, "No se encontró la hoja RUTAS"
    ws = wb["RUTAS"]
    ruta_camion = {}; ruta_dia = {}; ruta_fecha = {}
    agente_camion = {}; agente_dia = {}
    camion_ciudades = defaultdict(list)

    for i in range(3, ws.max_row+1):
        camion = str(ws.cell(i,1).value or "").strip()
        ciudad = norm(str(ws.cell(i,2).value or ""))
        agente = str(ws.cell(i,3).value or "").strip()
        dia    = str(ws.cell(i,4).value or "").strip().upper()
        v5     = ws.cell(i,5).value

        if not camion and not ciudad: continue

        if isinstance(v5, (datetime.datetime, datetime.date)):
            fecha = v5.strftime("%d/%m/%Y")
        elif v5:
            fecha = str(v5).strip()
        else:
            fecha = "--"

        if ciudad and ciudad not in ruta_camion:
            ruta_camion[ciudad] = camion
            ruta_dia[ciudad]    = dia
            ruta_fecha[ciudad]  = fecha

        if agente and agente not in agente_camion:
            agente_camion[agente] = camion
            agente_dia[agente]    = dia

        if camion and ciudad:
            if ciudad not in camion_ciudades[camion]:
                camion_ciudades[camion].append(ciudad)

    return ruta_camion, ruta_dia, ruta_fecha, agente_camion, agente_dia, dict(camion_ciudades), None

def procesar_data(wb, centros_nom, centros_emp, cli_vend, cli_prio,
                  agente_camion, agente_dia):
    """
    Lee DATA desde fila 3 (cabecera en fila 2).
    Agrupa por ciudad Y por camión.
    Registra clientes no encontrados y sin camión.
    """
    if "DATA" not in wb.sheetnames:
        return None, "No se encontró la hoja DATA"
    ws = wb["DATA"]
    ci = detectar_cols(ws, hdr_row=2)
    last = ws.max_row
    if last < 3:
        return None, "No hay datos en DATA desde fila 3."

    # Por destino
    df_dest   = defaultdict(list)
    df_monto  = defaultdict(float)
    df_fecha  = {}
    df_ruta   = {}
    df_bodega = {}
    df_emp    = defaultdict(set)

    # Por camión
    df_camion      = defaultdict(list)
    df_cam_monto   = defaultdict(float)
    df_cam_bodega  = defaultdict(set)
    df_cam_emp     = defaultdict(set)
    df_cam_ciudades= defaultdict(set)

    no_encontrados = {}   # cod -> {nombre, ciudad, comps, monto}
    sin_camion     = {}   # cod -> {nombre, ciudad, agente, comps, monto}

    for i in range(3, last+1):
        ciudad = norm(str(ws.cell(i, ci["ciudad"]).value or ""))
        if not ciudad: continue

        comp_str = str(ws.cell(i, ci["comp"]).value or "").strip()
        try:    monto = float(ws.cell(i, ci["total"]).value or 0)
        except: monto = 0.0

        bod_v = ws.cell(i, ci["bodega"]).value
        try:    bodega = str(int(float(str(bod_v))))
        except: bodega = str(bod_v or "").strip()

        emp_res   = centros_emp.get(bodega, "")
        centro_nom = centros_nom.get(bodega, "")

        cod_cli  = norm(str(ws.cell(i, ci["codcli"]).value or ""))
        nom_dest = str(ws.cell(i, ci["nomdest"]).value or "").strip()
        ruta_str = str(ws.cell(i, ci["ruta"]).value or "").strip()

        serial, fecha_txt = fecha_celda(ws.cell(i, ci["fecha"]))

        # Buscar agente en CLIENTES
        if cod_cli in cli_vend:
            agente   = cli_vend[cod_cli]
            prioridad = cli_prio[cod_cli]
        else:
            agente    = "SIN AGENTE"
            prioridad = 9999
            if cod_cli not in no_encontrados:
                no_encontrados[cod_cli] = {
                    "nombre": nom_dest, "ciudad": ciudad,
                    "comps": [], "monto": 0.0
                }
            no_encontrados[cod_cli]["comps"].append(comp_str)
            no_encontrados[cod_cli]["monto"] += monto

        # Buscar camión via agente
        camion = agente_camion.get(agente, "")
        dia_cam = agente_dia.get(agente, "--")

        if not camion and agente != "SIN AGENTE":
            if cod_cli not in sin_camion:
                sin_camion[cod_cli] = {
                    "nombre": nom_dest, "ciudad": ciudad,
                    "agente": agente, "comps": [], "monto": 0.0
                }
            sin_camion[cod_cli]["comps"].append(comp_str)
            sin_camion[cod_cli]["monto"] += monto

        row = {
            "comp": comp_str, "nom_dest": nom_dest,
            "monto": monto, "serial": serial, "fecha_txt": fecha_txt,
            "cod_cli": cod_cli, "bodega": bodega,
            "ciudad": ciudad, "agente": agente,
            "prioridad": prioridad, "centro_nom": centro_nom,
            "camion": camion or "SIN CAMION",
            "dia_cam": dia_cam, "ruta_str": ruta_str,
        }

        # Agrupacion por destino
        df_dest[ciudad].append(row)
        df_monto[ciudad] += monto
        if emp_res: df_emp[ciudad].add(emp_res)
        if ciudad not in df_fecha:
            df_fecha[ciudad]  = fecha_txt
            df_ruta[ciudad]   = ruta_str
            df_bodega[ciudad] = bodega

        # Agrupacion por camión
        cam_key = camion if camion else "SIN CAMION"
        df_camion[cam_key].append(row)
        df_cam_monto[cam_key]  += monto
        if emp_res: df_cam_emp[cam_key].add(emp_res)
        df_cam_ciudades[cam_key].add(ciudad)
        df_cam_bodega[cam_key].add(bodega)

    return {
        # Por destino
        "df_dest": df_dest, "df_monto": df_monto,
        "df_fecha": df_fecha, "df_ruta": df_ruta,
        "df_bodega": df_bodega, "df_emp": df_emp,
        # Por camión
        "df_camion": df_camion, "df_cam_monto": df_cam_monto,
        "df_cam_emp": df_cam_emp, "df_cam_ciudades": df_cam_ciudades,
        "df_cam_bodega": df_cam_bodega,
        # Reportes de calidad
        "no_encontrados": no_encontrados,
        "sin_camion": sin_camion,
    }, None

# ── Escribir hoja de ruta ─────────────────────────────────────
ANCHOS = {"A":11.43,"B":7.43,"C":6.86,"D":11.86,"E":39.14,
          "F":16.43,"G":11.0,"H":8.86,"I":6.0,"J":31.29}

HDRS = ["Seq.","Prior.","Bodega","Comprobante","Nombre Cliente",
        "Agente","Fecha Fact.","Total ($)","Bultos","Observaciones"]

def escribir_hoja(wb_out, ws_name, titulo_cab, subtitulo_cab,
                  registros, bodega_str, empresas, fecha_str,
                  centros_nom, modo_extra="", separar_ciudades=False):
    """Escribe una hoja de ruta genérica (sirve para destino y camión)"""
    wsR = wb_out.create_sheet(ws_name)

    for col, w in ANCHOS.items():
        wsR.column_dimensions[col].width = w

    # F1 Titulo
    wsR.merge_cells("A1:J1")
    c = wsR.cell(1,1)
    c.value = "RUTERO DE DESPACHO"
    c.font  = Font(name="Arial", bold=True, size=13, color=C_BLN)
    c.fill  = PatternFill("solid", fgColor=C_AZO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin()
    wsR.row_dimensions[1].height = 21.95

    # F2: EMPRESA
    lbl(wsR, 2, 1, 2, "EMPRESA:")
    val(wsR, 2, 3, 10, empresas or "--", bold=True)

    # F3: titulo izq | fecha der
    lbl(wsR, 3, 1, 2, titulo_cab + ":")
    val(wsR, 3, 3, 5, subtitulo_cab)
    lbl(wsR, 3, 6, 7, "FECHA FACTURA:")
    val(wsR, 3, 8, 10, fecha_str)

    # F4: BODEGA
    lbl(wsR, 4, 1, 2, "BODEGA / CENTRO:")
    val(wsR, 4, 3, 10, bodega_str)

    # F5: FECHA ENVIO | CHOFER | PLACA
    lbl(wsR, 5, 1, 1, "FECHA ENVIO:")
    linea(wsR, 5, 2, 3)
    lbl(wsR, 5, 4, 4, "CHOFER:")
    linea(wsR, 5, 5, 7)
    lbl(wsR, 5, 8, 8, "PLACA:")
    linea(wsR, 5, 9, 10)

    n_rows = len(registros)
    monto_t = sum(r["monto"] for r in registros)

    # F6: N FACTURAS | MONTO
    lbl(wsR, 6, 1, 2, "N\u00b0 FACTURAS:")
    val(wsR, 6, 3, 3, n_rows, bold=True)
    lbl(wsR, 6, 4, 5, "MONTO TOTAL:")
    val(wsR, 6, 6, 8, round(monto_t, 2), bold=True, fmt="$#,##0.00")
    wsR.merge_cells("I6:J6")
    wsR.cell(6,9).fill   = PatternFill("solid", fgColor=C_BLN)
    wsR.cell(6,9).border = thin()

    for rr in range(2, 7):
        wsR.row_dimensions[rr].height = 15.0

    # Linea separadora
    for col in range(1, 11):
        c = wsR.cell(6, col)
        ob = c.border
        c.border = Border(left=ob.left, right=ob.right, top=ob.top,
                          bottom=Side(style="medium", color=C_AZO))

    # F7: Link volver
    wsR.merge_cells("A7:J7")
    c7 = wsR.cell(7,1)
    c7.value = "<< Volver al Resumen"
    c7.font  = Font(name="Arial", size=8, italic=True, color=C_AZM)
    c7.fill  = PatternFill("solid", fgColor=C_GRS)
    c7.alignment = Alignment(horizontal="left", vertical="center")
    c7.border = thin()
    c7.hyperlink = "#RESUMEN!A1"
    wsR.row_dimensions[7].height = 12.0

    # F8: Cabecera tabla
    for fc, h in enumerate(HDRS, 1):
        c = wsR.cell(8, fc)
        c.value = h
        c.font  = Font(name="Arial", bold=True, size=8, color=C_BLN)
        c.fill  = PatternFill("solid", fgColor=C_AZM)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
    wsR.row_dimensions[8].height = 15.95

    # Datos
    ds = 9
    dr = ds
    seq = 1
    ciudad_actual = None

    for row in registros:
        # Separador de ciudad (solo en modo camion)
        if separar_ciudades and row["ciudad"] != ciudad_actual:
            ciudad_actual = row["ciudad"]
            # Fila separadora con nombre de ciudad
            wsR.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=10)
            c = wsR.cell(dr, 1)
            c.value = f"  📍  {ciudad_actual}"
            c.font  = Font(name="Arial", bold=True, size=9, color=C_BLN)
            c.fill  = PatternFill("solid", fgColor=C_AZM)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = thin()
            wsR.row_dimensions[dr].height = 16
            dr += 1

        for fc in range(1, 11):
            c = wsR.cell(dr, fc)
            c.fill = PatternFill("solid", fgColor=C_BLN)
            c.font = Font(name="Arial", size=8, color="323232")
            c.border = thin()
            c.alignment = Alignment(
                horizontal="left" if fc == 5 else "center",
                vertical="center")

        wsR.cell(dr,1).value = seq
        wsR.cell(dr,2).value = row["prioridad"]
        wsR.cell(dr,3).value = row["bodega"]
        wsR.cell(dr,4).value = row["comp"]
        wsR.cell(dr,5).value = row["nom_dest"]

        wsR.cell(dr,6).value = row["agente"]
        if row["agente"] == "SIN AGENTE":
            wsR.cell(dr,6).font = Font(name="Arial",size=8,color=C_RJO,italic=True)

        s = row["serial"]
        if s and s > 1:
            wsR.cell(dr,7).value = s
            wsR.cell(dr,7).number_format = "DD/MM/YYYY"
        else:
            wsR.cell(dr,7).value = row["fecha_txt"]

        wsR.cell(dr,8).value = round(row["monto"], 2)
        wsR.cell(dr,8).number_format = "#,##0.00"
        wsR.cell(dr,10).alignment = Alignment(horizontal="left",vertical="center")
        wsR.row_dimensions[dr].height = 14.1

        p = row["prioridad"]
        if p == 1:
            wsR.cell(dr,2).font = Font(name="Arial",size=8,color=C_RJO,bold=True)
        elif p == 2:
            wsR.cell(dr,2).font = Font(name="Arial",size=8,color="823C00",bold=True)
        else:
            wsR.cell(dr,2).font = Font(name="Arial",size=8,color="646464")

        dr += 1
        seq += 1

    # Subtotal — dr apunta a la siguiente fila libre despues de datos
    sub_r = dr
    wsR.merge_cells(start_row=sub_r,start_column=1,end_row=sub_r,end_column=7)
    c = wsR.cell(sub_r,1)
    c.value="SUBTOTAL"; c.font=Font(name="Arial",bold=True,size=9,color=C_BLN)
    c.fill=PatternFill("solid",fgColor=C_AZO)
    c.alignment=Alignment(horizontal="right",vertical="center"); c.border=thin()

    c=wsR.cell(sub_r,8)
    c.value=f"=SUM(H{ds}:H{ds+n_rows-1})"
    c.number_format="$#,##0.00"
    c.font=Font(name="Arial",bold=True,size=9,color=C_BLN)
    c.fill=PatternFill("solid",fgColor=C_AZO)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin()
    for fc in [9,10]:
        wsR.cell(sub_r,fc).fill=PatternFill("solid",fgColor=C_AZO)
        wsR.cell(sub_r,fc).border=thin()
    wsR.row_dimensions[sub_r].height=16

    # Firmas
    fir_r = sub_r+5
    wsR.merge_cells(start_row=fir_r,start_column=1,end_row=fir_r,end_column=3)
    c=wsR.cell(fir_r,1); c.value="Firma Chofer: _______________________"
    c.font=Font(name="Arial",size=9)
    c.alignment=Alignment(horizontal="left",vertical="center")
    wsR.merge_cells(start_row=fir_r,start_column=8,end_row=fir_r,end_column=10)
    c=wsR.cell(fir_r,8); c.value="Firma Receptor: _______________________"
    c.font=Font(name="Arial",size=9)
    c.alignment=Alignment(horizontal="right",vertical="center")
    wsR.row_dimensions[fir_r].height=20

    # Impresión
    wsR.page_setup.orientation="landscape"; wsR.page_setup.paperSize=9
    wsR.page_setup.fitToWidth=1; wsR.page_setup.fitToHeight=0
    wsR.sheet_properties.pageSetUpPr.fitToPage=True
    wsR.page_margins.top=0.75; wsR.page_margins.bottom=0.75
    wsR.page_margins.left=0.25; wsR.page_margins.right=0.25
    wsR.print_title_rows="1:8"
    wsR.print_area=f"A1:J{fir_r}"
    wsR.oddFooter.center.text="&\"Arial\"&8 Página &P de &N"

    return n_rows, monto_t

# ── Hoja resumen ──────────────────────────────────────────────
def escribir_resumen(ws_res, filas_res, total_f, total_m, modo):
    ws_res.merge_cells("A1:H1")
    c=ws_res.cell(1,1)
    c.value=f"RESUMEN DE DESPACHOS — {'POR DESTINO' if modo=='destino' else 'POR CAMIÓN'}"
    c.font=Font(name="Arial",bold=True,size=13,color=C_BLN)
    c.fill=PatternFill("solid",fgColor=C_AZO)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.border=thin()
    ws_res.row_dimensions[1].height=28

    ws_res.merge_cells("A2:H2")
    c=ws_res.cell(2,1)
    c.value=(f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
             f"   |   Facturas: {total_f}   |   Monto: ${total_m:,.2f}")
    c.font=Font(name="Arial",size=9,italic=True)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws_res.row_dimensions[2].height=16

    hdrs=["AGRUPACIÓN","N° FACT.","MONTO ($)","EMPRESAS","HOJA"]
    widths=[30,12,16,40,10]
    for i,(h,w) in enumerate(zip(hdrs,widths),1):
        c=ws_res.cell(4,i)
        c.value=h; c.font=Font(name="Arial",bold=True,size=10,color=C_BLN)
        c.fill=PatternFill("solid",fgColor=C_AZM)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin()
        ws_res.column_dimensions[get_column_letter(i)].width=w
    ws_res.row_dimensions[4].height=22

    for row_num, (nombre, n, m, emp, sn) in enumerate(filas_res, 5):
        bg = C_AZC if row_num%2==0 else C_BLN
        for fc,v in enumerate([nombre, n, round(m,2), emp], 1):
            c=ws_res.cell(row_num,fc)
            c.value=v; c.font=Font(name="Arial",size=9)
            c.fill=PatternFill("solid",fgColor=bg); c.border=thin()
            c.alignment=Alignment(
                horizontal="left" if fc in [1,4] else "center",
                vertical="center")
            if fc==3: c.number_format="#,##0.00"
        lk=ws_res.cell(row_num,5)
        lk.value=">> Ver"; lk.font=Font(name="Arial",size=9,bold=True,color=C_AZM)
        lk.fill=PatternFill("solid",fgColor=bg); lk.border=thin()
        lk.alignment=Alignment(horizontal="center",vertical="center")
        lk.hyperlink=f"#{sn}!A1"
        ws_res.row_dimensions[row_num].height=16

    tot_r=len(filas_res)+5
    # Total general: estilizar SIN merge para evitar MergedCell read-only
    vals_tot={1:"TOTAL GENERAL", 2:"", 3:total_f,
              4:round(total_m,2), 5:""}
    for fc in range(1,6):
        c=ws_res.cell(tot_r,fc)
        c.value=vals_tot[fc]
        c.fill=PatternFill("solid",fgColor=C_AZO)
        c.font=Font(name="Arial",bold=True,size=10,color=C_BLN)
        c.border=thin()
        c.alignment=Alignment(horizontal="center",vertical="center")
    ws_res.cell(tot_r,4).number_format="#,##0.00"
    ws_res.row_dimensions[tot_r].height=20

# ── Hojas de calidad ──────────────────────────────────────────
def hoja_no_encontrados(wb_out, no_enc):
    ws=wb_out.create_sheet("CLIENTES NO ENCONTRADOS")
    ws.sheet_properties.tabColor="ED7D31"
    ws.merge_cells("A1:F1")
    c=ws.cell(1,1); c.value="⚠  CLIENTES NO ENCONTRADOS EN HOJA CLIENTES"
    c.font=Font(name="Arial",bold=True,size=12,color=C_BLN)
    c.fill=PatternFill("solid",fgColor="ED7D31")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    c.border=thin(); ws.row_dimensions[1].height=26

    ws.merge_cells("A2:F2")
    c=ws.cell(2,1)
    c.value="Estos clientes aparecen en DATA pero no tienen registro en CLIENTES. Agréguelos para que aparezcan correctamente en el rutero."
    c.font=Font(name="Arial",italic=True,size=9,color="7F3F00")
    c.fill=PatternFill("solid",fgColor=C_AMR)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True)
    c.border=thin(); ws.row_dimensions[2].height=28

    hdrs=[("CÓDIGO CLIENTE",20),("NOMBRE",36),("CIUDAD",22),
          ("N° COMPROBANTES",18),("MONTO TOTAL",16),("ACCIÓN",20)]
    for i,(h,w) in enumerate(hdrs,1):
        c=ws.cell(3,i); c.value=h
        c.font=Font(name="Arial",bold=True,size=10,color=C_BLN)
        c.fill=PatternFill("solid",fgColor="ED7D31")
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=thin()
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[3].height=22

    if not no_enc:
        ws.merge_cells("A4:F4")
        c=ws.cell(4,1); c.value="✅  Sin pendientes — todos los clientes están registrados"
        c.font=Font(name="Arial",italic=True,size=10,color=C_VRD)
        c.fill=PatternFill("solid",fgColor=C_MNT)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=thin(); ws.row_dimensions[4].height=22
        return

    for idx,(cod,info) in enumerate(sorted(no_enc.items()), 4):
        bg=C_SLM if idx%2==0 else "FFF0E8"
        vals=[cod, info["nombre"], info["ciudad"],
              len(info["comps"]), round(info["monto"],2), "Agregar a CLIENTES"]
        for fc,v in enumerate(vals,1):
            c=ws.cell(idx,fc); c.value=v
            c.font=Font(name="Arial",size=9,
                       color=C_RJO if fc==6 else "323232",
                       bold=(fc==6))
            c.fill=PatternFill("solid",fgColor=bg); c.border=thin()
            c.alignment=Alignment(
                horizontal="left" if fc in [2,3,6] else "center",
                vertical="center")
            if fc==5: c.number_format="#,##0.00"
        ws.row_dimensions[idx].height=18
    ws.freeze_panes="A4"

def hoja_duplicados(wb_out, dupl):
    ws=wb_out.create_sheet("CLIENTES DUPLICADOS")
    ws.sheet_properties.tabColor="7030A0"
    ws.merge_cells("A1:E1")
    c=ws.cell(1,1); c.value="⚠  CLIENTES DUPLICADOS EN HOJA CLIENTES"
    c.font=Font(name="Arial",bold=True,size=12,color=C_BLN)
    c.fill=PatternFill("solid",fgColor="7030A0")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    c.border=thin(); ws.row_dimensions[1].height=26

    ws.merge_cells("A2:E2")
    c=ws.cell(2,1)
    c.value="Estos códigos aparecen más de una vez en la hoja CLIENTES. Solo se usa el primer registro. Revise y corrija."
    c.font=Font(name="Arial",italic=True,size=9,color="4B0082")
    c.fill=PatternFill("solid",fgColor="E8E0F0")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True)
    c.border=thin(); ws.row_dimensions[2].height=28

    hdrs=[("CÓDIGO CLIENTE",20),("NOMBRE",36),("AGENTE",16),
          ("PRIORIDAD",14),("FILA EN CLIENTES",18)]
    for i,(h,w) in enumerate(hdrs,1):
        c=ws.cell(3,i); c.value=h
        c.font=Font(name="Arial",bold=True,size=10,color=C_BLN)
        c.fill=PatternFill("solid",fgColor="7030A0")
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=thin()
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[3].height=22

    if not dupl:
        ws.merge_cells("A4:E4")
        c=ws.cell(4,1); c.value="✅  Sin duplicados — base de clientes limpia"
        c.font=Font(name="Arial",italic=True,size=10,color=C_VRD)
        c.fill=PatternFill("solid",fgColor=C_MNT)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=thin(); ws.row_dimensions[4].height=22
        return

    for idx,d in enumerate(dupl,4):
        bg="F3E8FF" if idx%2==0 else "EDD9FF"
        for fc,v in enumerate([d["codigo"],d["nombre"],d["agente"],
                                d["prioridad"],d["fila"]],1):
            c=ws.cell(idx,fc); c.value=v
            c.font=Font(name="Arial",size=9)
            c.fill=PatternFill("solid",fgColor=bg); c.border=thin()
            c.alignment=Alignment(
                horizontal="left" if fc==2 else "center",
                vertical="center")
        ws.row_dimensions[idx].height=18
    ws.freeze_panes="A4"

def hoja_sin_camion(wb_out, sin_cam):
    ws=wb_out.create_sheet("CLIENTES SIN CAMION")
    ws.sheet_properties.tabColor="4472C4"
    ws.merge_cells("A1:F1")
    c=ws.cell(1,1); c.value="ℹ  CLIENTES SIN CAMIÓN ASIGNADO"
    c.font=Font(name="Arial",bold=True,size=12,color=C_BLN)
    c.fill=PatternFill("solid",fgColor="4472C4")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    c.border=thin(); ws.row_dimensions[1].height=26

    ws.merge_cells("A2:F2")
    c=ws.cell(2,1)
    c.value="Estos clientes tienen agente asignado pero el agente NO está en la hoja RUTAS. Asigne el camión correspondiente."
    c.font=Font(name="Arial",italic=True,size=9,color="1F3864")
    c.fill=PatternFill("solid",fgColor=C_AZC)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True)
    c.border=thin(); ws.row_dimensions[2].height=28

    hdrs=[("CÓDIGO CLIENTE",20),("NOMBRE",36),("CIUDAD",22),
          ("AGENTE",14),("N° COMP.",14),("MONTO",16)]
    for i,(h,w) in enumerate(hdrs,1):
        c=ws.cell(3,i); c.value=h
        c.font=Font(name="Arial",bold=True,size=10,color=C_BLN)
        c.fill=PatternFill("solid",fgColor="4472C4")
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=thin()
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[3].height=22

    if not sin_cam:
        ws.merge_cells("A4:F4")
        c=ws.cell(4,1); c.value="✅  Todos los clientes con agente tienen camión asignado"
        c.font=Font(name="Arial",italic=True,size=10,color=C_VRD)
        c.fill=PatternFill("solid",fgColor=C_MNT)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=thin(); ws.row_dimensions[4].height=22
        return

    for idx,(cod,info) in enumerate(sorted(sin_cam.items()),4):
        bg=C_AZC if idx%2==0 else "EBF3FB"
        for fc,v in enumerate([cod,info["nombre"],info["ciudad"],
                                info["agente"],len(info["comps"]),
                                round(info["monto"],2)],1):
            c=ws.cell(idx,fc); c.value=v
            c.font=Font(name="Arial",size=9)
            c.fill=PatternFill("solid",fgColor=bg); c.border=thin()
            c.alignment=Alignment(
                horizontal="left" if fc in [2,3] else "center",
                vertical="center")
            if fc==6: c.number_format="#,##0.00"
        ws.row_dimensions[idx].height=18
    ws.freeze_panes="A4"

# ── Generar libro Por Destino ─────────────────────────────────
def generar_por_destino(datos, centros_nom, ruta_camion, ruta_dia, ruta_fecha):
    wb_out=openpyxl.Workbook(); wb_out.remove(wb_out.active)
    ws_res=wb_out.create_sheet("RESUMEN")

    ciudades=sorted(datos["df_dest"].keys())
    filas_res=[]; total_f=0; total_m=0.0

    for ciudad in ciudades:
        regs=datos["df_dest"][ciudad]
        regs.sort(key=lambda r:(r["prioridad"],
                                r["comp"].zfill(20) if r["comp"].isdigit() else r["comp"]))
        sn=sh_name(ciudad)
        bodega=datos["df_bodega"].get(ciudad,"")
        bodega_str=bodega
        if bodega in centros_nom and centros_nom[bodega]:
            bodega_str=bodega+" -- "+centros_nom[bodega]
        emp=", ".join(sorted(datos["df_emp"][ciudad])) or "--"
        fecha=datos["df_fecha"].get(ciudad,"--")

        n,m=escribir_hoja(wb_out, sn, "CIUDAD DEST.", ciudad,
                          regs, bodega_str, emp, fecha, centros_nom)
        total_f+=n; total_m+=m
        filas_res.append((ciudad, n, m, emp, sn))

    hoja_no_encontrados(wb_out, datos["no_encontrados"])
    hoja_duplicados(wb_out, datos.get("duplicados",[]))
    hoja_sin_camion(wb_out, datos["sin_camion"])
    escribir_resumen(ws_res, filas_res, total_f, total_m, "destino")
    return wb_out

# ── Generar libro Por Camión ──────────────────────────────────
def generar_por_camion(datos, centros_nom, agente_camion):
    wb_out=openpyxl.Workbook(); wb_out.remove(wb_out.active)
    ws_res=wb_out.create_sheet("RESUMEN")

    camiones=sorted(datos["df_camion"].keys())
    filas_res=[]; total_f=0; total_m=0.0

    for camion in camiones:
        regs=datos["df_camion"][camion]
        regs.sort(key=lambda r:(
            r["ciudad"],
            r["prioridad"],
            r["comp"].zfill(20) if r["comp"].isdigit() else r["comp"]
        ))
        sn=sh_name(camion)
        ciudades_cam=", ".join(sorted(datos["df_cam_ciudades"][camion]))
        bodegas=datos["df_cam_bodega"][camion]
        bodega_str="; ".join(
            (b+" -- "+centros_nom[b] if b in centros_nom else b)
            for b in sorted(bodegas)
        )
        emp=", ".join(sorted(datos["df_cam_emp"][camion])) or "--"
        fecha=regs[0]["fecha_txt"] if regs else "--"

        n,m=escribir_hoja(wb_out, sn, "CAMIÓN", camion,
                          regs, bodega_str, emp, fecha, centros_nom,
                          modo_extra=f"Ciudades: {ciudades_cam}",
                          separar_ciudades=True)
        total_f+=n; total_m+=m
        filas_res.append((camion, n, m, emp, sn))

    hoja_no_encontrados(wb_out, datos["no_encontrados"])
    hoja_duplicados(wb_out, datos.get("duplicados",[]))
    hoja_sin_camion(wb_out, datos["sin_camion"])
    escribir_resumen(ws_res, filas_res, total_f, total_m, "camion")
    return wb_out

# ── Interfaz Streamlit ────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Rutero de Despacho",
        page_icon="🚚", layout="centered"
    )

    st.markdown("""
        <div style="background:#2C5F8A;padding:18px 24px;border-radius:8px;margin-bottom:24px">
            <h2 style="color:white;margin:0;font-family:Arial">🚚 Generador de Rutero de Despacho</h2>
            <p style="color:#DEEAF7;margin:6px 0 0 0;font-size:14px">
                Suba la plantilla y seleccione el tipo de rutero a generar
            </p>
        </div>
    """, unsafe_allow_html=True)

    uploaded=st.file_uploader(
        "Seleccione la plantilla (.xlsx o .xlsm)",
        type=["xlsx","xlsm"],
        help="Debe contener las hojas: DATA, CENTROS, CLIENTES, RUTAS"
    )

    if not uploaded:
        st.info("📂 Suba la plantilla para continuar.")
        with st.expander("📋 Estructura requerida"):
            st.markdown("""
            | Hoja | Cabecera | Datos desde |
            |---|---|---|
            | DATA | Fila 2 | Fila 3 |
            | CENTROS | Fila 2 | Fila 3 |
            | CLIENTES | Fila 2 | Fila 3 |
            | RUTAS | Fila 2 | Fila 3 |
            """)
        return

    # Modo de rutero
    st.markdown("### Tipo de Rutero")
    modo=st.radio(
        "Seleccione cómo generar el rutero:",
        ["📍 Por Destino (una hoja por ciudad)",
         "🚛 Por Camión (una hoja por camión)"],
        horizontal=True
    )
    modo_key="destino" if "Destino" in modo else "camion"

    if st.button("🚀 Generar Rutero", type="primary", use_container_width=True):
        with st.spinner("Procesando plantilla..."):
            try:
                wb=openpyxl.load_workbook(uploaded, keep_vba=False, data_only=True)

                hojas_req=["DATA","CENTROS","CLIENTES","RUTAS"]
                faltantes=[h for h in hojas_req if h not in wb.sheetnames]
                if faltantes:
                    st.error(f"Faltan hojas: {', '.join(faltantes)}"); return

                with st.status("Cargando maestros...") as status:
                    centros_nom, centros_emp, err=cargar_centros(wb)
                    if err: st.error(err); return

                    cli_vend, cli_prio, duplicados, err=cargar_clientes(wb)
                    if err: st.error(err); return

                    (ruta_camion, ruta_dia, ruta_fecha,
                     agente_camion, agente_dia,
                     camion_ciudades, err)=cargar_rutas(wb)
                    if err: st.error(err); return

                    status.update(label="Procesando DATA...")
                    datos, err=procesar_data(
                        wb, centros_nom, centros_emp,
                        cli_vend, cli_prio,
                        agente_camion, agente_dia
                    )
                    if err: st.error(err); return
                    datos["duplicados"]=duplicados

                    status.update(label="Generando Excel...")
                    if modo_key=="destino":
                        wb_out=generar_por_destino(
                            datos, centros_nom, ruta_camion, ruta_dia, ruta_fecha)
                    else:
                        wb_out=generar_por_camion(datos, centros_nom, agente_camion)

                    status.update(label="✅ Listo", state="complete")

                # Métricas
                no_enc=len(datos["no_encontrados"])
                sin_cam=len(datos["sin_camion"])
                dupl=len(datos["duplicados"])

                if modo_key=="destino":
                    n_hojas=len(datos["df_dest"])
                    total_f=sum(len(v) for v in datos["df_dest"].values())
                    total_m=sum(datos["df_monto"].values())
                else:
                    n_hojas=len(datos["df_camion"])
                    total_f=sum(len(v) for v in datos["df_camion"].values())
                    total_m=sum(datos["df_cam_monto"].values())

                col1,col2,col3=st.columns(3)
                col1.metric("Hojas generadas", n_hojas)
                col2.metric("Total facturas", f"{total_f:,}")
                col3.metric("Monto total", f"${total_m:,.2f}")

                if no_enc or sin_cam or dupl:
                    st.markdown("#### ⚠️ Alertas de calidad")
                    c1,c2,c3=st.columns(3)
                    if no_enc:
                        c1.warning(f"**{no_enc}** clientes sin registro\n\nVer hoja *CLIENTES NO ENCONTRADOS*")
                    if sin_cam:
                        c2.warning(f"**{sin_cam}** clientes sin camión\n\nVer hoja *CLIENTES SIN CAMION*")
                    if dupl:
                        c3.warning(f"**{dupl}** registros duplicados\n\nVer hoja *CLIENTES DUPLICADOS*")

                buf=io.BytesIO()
                wb_out.save(buf); buf.seek(0)
                tipo="DESTINO" if modo_key=="destino" else "CAMION"
                nombre=f"RUTERO_{tipo}_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

                st.download_button(
                    label="📥 Descargar Rutero Excel",
                    data=buf, file_name=nombre,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary"
                )

            except Exception as e:
                st.error(f"Error inesperado: {str(e)}")
                st.exception(e)

if __name__=="__main__":
    main()
